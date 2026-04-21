#!/usr/bin/env python3
from __future__ import annotations

import argparse
import html
import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


CATEGORY_COLORS = {
    "市场": {"accent": "#9c5f2f", "soft": "#f1ddc9"},
    "产品": {"accent": "#1f5e5b", "soft": "#d2ece8"},
    "研究报告": {"accent": "#7b3f73", "soft": "#f0d7eb"},
    "研究": {"accent": "#5e4e2d", "soft": "#ebdfbb"},
    "产品/研究": {"accent": "#2f4a8b", "soft": "#d9e3fb"},
    "未分类": {"accent": "#51606b", "soft": "#dde4e9"},
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\r", "\n")
    lines = [line.strip() for line in text.split("\n")]
    return "\n".join(line for line in lines if line)


def parse_time_label(label: str) -> dict[str, Any]:
    normalized = re.sub(r"\s+", "", label)
    year_match = re.search(r"(\d{4})年", normalized)
    if not year_match:
        raise ValueError(f"无法解析年份: {label!r}")

    year = int(year_match.group(1))
    range_match = re.search(r"(\d{1,2})-(\d{1,2})月", normalized)
    if range_match:
        start_month = int(range_match.group(1))
        end_month = int(range_match.group(2))
    else:
        month_match = re.search(r"(\d{1,2})月", normalized)
        if month_match:
            start_month = end_month = int(month_match.group(1))
        else:
            start_month = end_month = 1

    date_key = f"{year:04d}-{start_month:02d}"
    return {
        "label": clean_text(label),
        "year": year,
        "start_month": start_month,
        "end_month": end_month,
        "date_key": date_key,
    }


def extract_links(raw: str) -> list[str]:
    return re.findall(r"https?://[^\s]+", raw)


def build_event(row: tuple[Any, ...], index: int) -> dict[str, Any]:
    time_value, category_value, event_value, significance_value, link_value = row
    time_meta = parse_time_label(clean_text(time_value))
    category = clean_text(category_value) or "未分类"
    if category not in CATEGORY_COLORS:
        CATEGORY_COLORS[category] = {"accent": "#51606b", "soft": "#dde4e9"}

    event = clean_text(event_value)
    significance = clean_text(significance_value)
    links = extract_links(clean_text(link_value))
    headline = event.split("。", 1)[0].strip() if event else time_meta["label"]
    return {
        "index": index,
        "time": time_meta["label"],
        "year": time_meta["year"],
        "start_month": time_meta["start_month"],
        "end_month": time_meta["end_month"],
        "date_key": time_meta["date_key"],
        "category": category,
        "event": event,
        "headline": headline,
        "significance": significance,
        "links": links,
    }


def load_events(workbook_path: Path) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook.active
    events = []
    for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
        if not any(row):
            continue
        events.append(build_event(row, index))
    events.sort(key=lambda item: (item["year"], item["start_month"], item["end_month"], item["index"]))
    return events


def category_badges(category_counts: Counter[str]) -> str:
    badges = []
    for category, count in sorted(category_counts.items(), key=lambda item: (-item[1], item[0])):
        palette = CATEGORY_COLORS[category]
        badges.append(
            f"""
            <button class="legend-chip filter-chip is-active" type="button" data-filter="{html.escape(category)}"
              style="--chip-accent: {palette['accent']}; --chip-soft: {palette['soft']};">
              <span class="legend-dot"></span>
              <span>{html.escape(category)}</span>
              <strong>{count}</strong>
            </button>
            """
        )
    return "\n".join(badges)


def year_nav(year_counts: Counter[int]) -> str:
    items = []
    for year in sorted(year_counts):
        items.append(
            f"""
            <a class="year-pill" href="#year-{year}">
              <span>{year}</span>
              <strong>{year_counts[year]}</strong>
            </a>
            """
        )
    return "\n".join(items)


def event_cards(events: list[dict[str, Any]]) -> str:
    pieces: list[str] = []
    current_year: int | None = None

    for offset, item in enumerate(events):
        if item["year"] != current_year:
            current_year = item["year"]
            pieces.append(
                f"""
                <section class="year-block" id="year-{current_year}">
                  <div class="year-header">
                    <div>
                      <p class="eyebrow">Year Marker</p>
                      <h2>{current_year}</h2>
                    </div>
                    <p>{sum(1 for event in events if event["year"] == current_year)} 个里程碑</p>
                  </div>
                """
            )

        palette = CATEGORY_COLORS[item["category"]]
        alignment = "left" if offset % 2 == 0 else "right"
        links_html = "".join(
            f'<a href="{html.escape(link)}" target="_blank" rel="noreferrer">{html.escape(link)}</a>'
            for link in item["links"]
        )
        range_note = (
            f'<span class="range-note">{item["start_month"]}-{item["end_month"]} 月</span>'
            if item["end_month"] != item["start_month"]
            else ""
        )
        pieces.append(
            f"""
            <article class="timeline-card {alignment}"
              data-category="{html.escape(item['category'])}"
              data-year="{item['year']}"
              data-search="{html.escape((item['headline'] + ' ' + item['event'] + ' ' + item['significance']).lower())}"
              style="--accent: {palette['accent']}; --soft: {palette['soft']};">
              <div class="timeline-node" aria-hidden="true"></div>
              <div class="card-shell">
                <div class="card-topline">
                  <span class="date-tag">{html.escape(item['time'])}</span>
                  {range_note}
                  <span class="category-tag">{html.escape(item['category'])}</span>
                </div>
                <h3>{html.escape(item['headline'])}</h3>
                <p class="event-copy">{html.escape(item['event'])}</p>
                <div class="meaning-block">
                  <h4>战略意义</h4>
                  <p>{html.escape(item['significance'])}</p>
                </div>
                <div class="links-block">{links_html}</div>
              </div>
            </article>
            """
        )

        is_last = offset == len(events) - 1 or events[offset + 1]["year"] != current_year
        if is_last:
            pieces.append("</section>")

    return "\n".join(pieces)


def render_html(events: list[dict[str, Any]], source_path: Path) -> str:
    generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    year_counts = Counter(item["year"] for item in events)
    category_counts = Counter(item["category"] for item in events)
    first_event = events[0]
    last_event = events[-1]
    stats_json = html.escape(json.dumps(events, ensure_ascii=False))

    return f"""<!DOCTYPE html>
<html lang="zh-CN">
<head>
  <meta charset="UTF-8">
  <meta name="viewport" content="width=device-width, initial-scale=1.0">
  <title>EigenPhi Milestones Timeline</title>
  <link rel="preconnect" href="https://fonts.googleapis.com">
  <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
  <link href="https://fonts.googleapis.com/css2?family=DM+Serif+Display:ital@0;1&family=IBM+Plex+Sans:wght@400;500;600;700&family=IBM+Plex+Mono:wght@400;500&display=swap" rel="stylesheet">
  <style>
    :root {{
      --bg: #f7f1e8;
      --bg-strong: #f0e6d9;
      --surface: rgba(255, 251, 245, 0.88);
      --surface-strong: rgba(255, 248, 239, 0.96);
      --ink: #1f1a17;
      --muted: #63574d;
      --line: rgba(56, 41, 28, 0.18);
      --shadow: 0 24px 70px rgba(71, 45, 18, 0.12);
      --hero: linear-gradient(135deg, rgba(161, 109, 49, 0.16), rgba(72, 89, 105, 0.08));
    }}

    * {{
      box-sizing: border-box;
    }}

    html {{
      scroll-behavior: smooth;
    }}

    body {{
      margin: 0;
      font-family: "IBM Plex Sans", "PingFang SC", "Hiragino Sans GB", sans-serif;
      color: var(--ink);
      background:
        radial-gradient(circle at top left, rgba(209, 183, 153, 0.35), transparent 28%),
        radial-gradient(circle at top right, rgba(111, 132, 155, 0.16), transparent 22%),
        linear-gradient(180deg, #f9f4ec 0%, #f4ecdf 100%);
      min-height: 100vh;
    }}

    body::before {{
      content: "";
      position: fixed;
      inset: 0;
      pointer-events: none;
      background-image:
        linear-gradient(rgba(79, 56, 38, 0.03) 1px, transparent 1px),
        linear-gradient(90deg, rgba(79, 56, 38, 0.03) 1px, transparent 1px);
      background-size: 26px 26px;
      mask-image: linear-gradient(180deg, rgba(0, 0, 0, 0.28), transparent 78%);
    }}

    a {{
      color: inherit;
    }}

    .page {{
      width: min(1200px, calc(100vw - 40px));
      margin: 0 auto;
      padding: 40px 0 72px;
      position: relative;
    }}

    .hero {{
      position: relative;
      overflow: hidden;
      border: 1px solid rgba(90, 64, 36, 0.08);
      border-radius: 32px;
      padding: 34px;
      background: var(--hero), var(--surface);
      box-shadow: var(--shadow);
      backdrop-filter: blur(14px);
    }}

    .hero::after {{
      content: "";
      position: absolute;
      right: -120px;
      top: -110px;
      width: 280px;
      height: 280px;
      border-radius: 50%;
      background: radial-gradient(circle, rgba(166, 122, 65, 0.22), transparent 65%);
    }}

    .eyebrow {{
      margin: 0 0 10px;
      font-size: 12px;
      letter-spacing: 0.18em;
      text-transform: uppercase;
      color: #8b6d53;
      font-family: "IBM Plex Mono", monospace;
    }}

    h1, h2, h3, h4 {{
      margin: 0;
      font-weight: 400;
    }}

    h1 {{
      max-width: 12ch;
      font-family: "DM Serif Display", Georgia, serif;
      font-size: clamp(3rem, 6vw, 5.2rem);
      line-height: 0.94;
      letter-spacing: -0.03em;
    }}

    .hero-grid {{
      display: grid;
      grid-template-columns: minmax(0, 1.4fr) minmax(280px, 0.9fr);
      gap: 28px;
      align-items: end;
      position: relative;
      z-index: 1;
    }}

    .hero-copy p {{
      max-width: 64ch;
      margin: 18px 0 0;
      color: var(--muted);
      font-size: 1rem;
      line-height: 1.7;
    }}

    .hero-meta {{
      display: grid;
      gap: 12px;
      justify-items: end;
    }}

    .meta-card {{
      width: 100%;
      border-radius: 20px;
      padding: 18px 18px 16px;
      background: var(--surface-strong);
      border: 1px solid rgba(103, 76, 47, 0.08);
    }}

    .meta-card strong {{
      display: block;
      font-size: 2rem;
      font-family: "DM Serif Display", Georgia, serif;
    }}

    .meta-card span {{
      display: block;
      margin-top: 6px;
      color: var(--muted);
      font-size: 0.95rem;
    }}

    .info-grid {{
      display: grid;
      grid-template-columns: 1.1fr 0.9fr;
      gap: 20px;
      margin-top: 24px;
    }}

    .panel {{
      border-radius: 28px;
      padding: 24px;
      background: var(--surface);
      border: 1px solid rgba(89, 61, 33, 0.08);
      box-shadow: 0 18px 50px rgba(81, 53, 26, 0.08);
      min-width: 0;
    }}

    .panel h2 {{
      font-family: "DM Serif Display", Georgia, serif;
      font-size: 2rem;
      margin-bottom: 14px;
    }}

    .stat-row {{
      display: grid;
      grid-template-columns: repeat(3, minmax(0, 1fr));
      gap: 12px;
    }}

    .stat-card {{
      border: 1px solid rgba(90, 64, 36, 0.08);
      border-radius: 20px;
      padding: 18px 16px;
      background: rgba(255, 250, 243, 0.72);
    }}

    .stat-card strong {{
      display: block;
      font-size: 2rem;
      font-family: "DM Serif Display", Georgia, serif;
    }}

    .stat-card span {{
      display: block;
      margin-top: 8px;
      color: var(--muted);
    }}

    .year-nav, .legend, .filter-bar {{
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
    }}

    .year-pill, .legend-chip {{
      display: inline-flex;
      align-items: center;
      gap: 10px;
      padding: 10px 14px;
      border-radius: 999px;
      text-decoration: none;
      background: rgba(255, 250, 244, 0.9);
      border: 1px solid rgba(93, 68, 40, 0.08);
      color: var(--ink);
      min-width: 0;
    }}

    .year-pill strong, .legend-chip strong {{
      font-family: "IBM Plex Mono", monospace;
      font-size: 0.85rem;
      color: var(--muted);
    }}

    .legend-chip {{
      background: var(--chip-soft, rgba(255, 250, 244, 0.9));
      border-color: color-mix(in srgb, var(--chip-accent, #8a5c31) 28%, white);
    }}

    .legend-dot {{
      width: 10px;
      height: 10px;
      border-radius: 50%;
      background: var(--chip-accent, #8a5c31);
      flex: 0 0 auto;
    }}

    .controls {{
      margin-top: 24px;
      display: grid;
      grid-template-columns: minmax(0, 1fr) auto;
      gap: 14px;
      align-items: center;
    }}

    .search-box {{
      display: flex;
      align-items: center;
      gap: 12px;
      padding: 14px 16px;
      border-radius: 18px;
      background: rgba(255, 251, 246, 0.9);
      border: 1px solid rgba(89, 61, 33, 0.08);
    }}

    .search-box input {{
      width: 100%;
      border: 0;
      outline: 0;
      background: transparent;
      color: var(--ink);
      font: inherit;
    }}

    .search-box span {{
      color: var(--muted);
      font-family: "IBM Plex Mono", monospace;
      font-size: 0.9rem;
    }}

    .filter-chip {{
      cursor: pointer;
      font: inherit;
    }}

    .filter-chip.is-active {{
      box-shadow: inset 0 0 0 1px color-mix(in srgb, var(--chip-accent, #8a5c31) 44%, white);
    }}

    .filter-chip.is-muted {{
      opacity: 0.42;
    }}

    .timeline-wrap {{
      position: relative;
      margin-top: 28px;
      padding: 10px 0 0;
    }}

    .timeline-wrap::before {{
      content: "";
      position: absolute;
      left: 50%;
      top: 0;
      bottom: 0;
      width: 2px;
      background:
        linear-gradient(180deg, rgba(153, 111, 60, 0.08) 0%, rgba(153, 111, 60, 0.5) 18%, rgba(31, 94, 91, 0.38) 55%, rgba(123, 63, 115, 0.44) 100%);
      transform: translateX(-50%);
      border-radius: 999px;
    }}

    .year-block {{
      position: relative;
      padding: 18px 0 26px;
    }}

    .year-header {{
      position: sticky;
      top: 12px;
      z-index: 3;
      width: fit-content;
      margin: 0 auto 18px;
      min-width: 280px;
      display: flex;
      align-items: end;
      justify-content: space-between;
      gap: 18px;
      padding: 18px 20px;
      border-radius: 22px;
      background: rgba(255, 249, 240, 0.92);
      border: 1px solid rgba(89, 61, 33, 0.08);
      box-shadow: 0 14px 36px rgba(81, 53, 26, 0.08);
      backdrop-filter: blur(10px);
    }}

    .year-header h2 {{
      font-size: 2.4rem;
      line-height: 0.9;
    }}

    .year-header p {{
      margin: 0;
      color: var(--muted);
      font-family: "IBM Plex Mono", monospace;
    }}

    .timeline-card {{
      position: relative;
      width: calc(50% - 42px);
      margin: 0 0 26px;
    }}

    .timeline-card.left {{
      margin-right: auto;
    }}

    .timeline-card.right {{
      margin-left: auto;
    }}

    .timeline-card[hidden] {{
      display: none;
    }}

    .timeline-node {{
      position: absolute;
      top: 30px;
      width: 18px;
      height: 18px;
      border-radius: 50%;
      background: var(--accent);
      border: 4px solid rgba(255, 251, 244, 0.95);
      box-shadow: 0 0 0 7px color-mix(in srgb, var(--accent) 18%, white);
      z-index: 2;
    }}

    .timeline-card.left .timeline-node {{
      right: -51px;
    }}

    .timeline-card.right .timeline-node {{
      left: -51px;
    }}

    .card-shell {{
      position: relative;
      border-radius: 26px;
      padding: 22px 22px 20px;
      background: linear-gradient(180deg, rgba(255, 252, 248, 0.96), rgba(254, 248, 240, 0.92));
      border: 1px solid rgba(84, 58, 28, 0.08);
      box-shadow: 0 20px 42px rgba(81, 53, 26, 0.08);
      min-width: 0;
    }}

    .card-shell::before {{
      content: "";
      position: absolute;
      inset: 0 auto 0 0;
      width: 5px;
      border-radius: 26px 0 0 26px;
      background: var(--accent);
    }}

    .card-topline {{
      display: flex;
      flex-wrap: wrap;
      gap: 8px;
      align-items: center;
      margin-bottom: 14px;
    }}

    .date-tag, .category-tag, .range-note {{
      display: inline-flex;
      align-items: center;
      gap: 6px;
      border-radius: 999px;
      padding: 7px 11px;
      font-size: 0.83rem;
      line-height: 1;
    }}

    .date-tag {{
      background: rgba(35, 27, 19, 0.06);
      font-family: "IBM Plex Mono", monospace;
    }}

    .category-tag {{
      background: var(--soft);
      color: color-mix(in srgb, var(--accent) 82%, black);
      font-weight: 600;
    }}

    .range-note {{
      border: 1px dashed color-mix(in srgb, var(--accent) 34%, white);
      color: color-mix(in srgb, var(--accent) 76%, black);
    }}

    .timeline-card h3 {{
      font-family: "DM Serif Display", Georgia, serif;
      font-size: 1.6rem;
      line-height: 1.1;
      margin-bottom: 12px;
    }}

    .event-copy, .meaning-block p {{
      margin: 0;
      color: var(--muted);
      line-height: 1.72;
      overflow-wrap: break-word;
    }}

    .meaning-block {{
      margin-top: 16px;
      padding: 16px;
      border-radius: 20px;
      background: color-mix(in srgb, var(--soft) 58%, white);
      border: 1px solid color-mix(in srgb, var(--accent) 12%, white);
    }}

    .meaning-block h4 {{
      margin-bottom: 8px;
      font-family: "IBM Plex Mono", monospace;
      font-size: 0.82rem;
      letter-spacing: 0.08em;
      text-transform: uppercase;
      color: color-mix(in srgb, var(--accent) 80%, black);
    }}

    .links-block {{
      margin-top: 16px;
      display: flex;
      flex-wrap: wrap;
      gap: 10px;
    }}

    .links-block a {{
      display: inline-flex;
      align-items: center;
      min-width: 0;
      max-width: 100%;
      padding: 10px 12px;
      border-radius: 14px;
      text-decoration: none;
      font-size: 0.9rem;
      background: rgba(255, 251, 246, 0.82);
      border: 1px solid rgba(93, 68, 40, 0.08);
      color: color-mix(in srgb, var(--accent) 80%, black);
      overflow-wrap: anywhere;
    }}

    .footnote {{
      margin-top: 26px;
      padding: 16px 18px;
      border-radius: 18px;
      background: rgba(255, 251, 246, 0.9);
      border: 1px solid rgba(89, 61, 33, 0.08);
      color: var(--muted);
      font-size: 0.92rem;
      line-height: 1.65;
    }}

    .empty-state {{
      display: none;
      margin-top: 24px;
      padding: 22px;
      border-radius: 20px;
      background: rgba(255, 251, 246, 0.9);
      border: 1px dashed rgba(89, 61, 33, 0.22);
      text-align: center;
      color: var(--muted);
    }}

    .empty-state.is-visible {{
      display: block;
    }}

    @media (max-width: 960px) {{
      .hero-grid, .info-grid, .controls {{
        grid-template-columns: 1fr;
      }}

      .hero-meta {{
        justify-items: stretch;
      }}

      .stat-row {{
        grid-template-columns: 1fr;
      }}

      .timeline-wrap::before {{
        left: 13px;
        transform: none;
      }}

      .year-header {{
        min-width: 0;
        width: 100%;
      }}

      .timeline-card {{
        width: calc(100% - 34px);
        margin-left: 34px;
      }}

      .timeline-card.right,
      .timeline-card.left {{
        margin-right: 0;
      }}

      .timeline-card .timeline-node {{
        left: -31px;
        right: auto;
      }}
    }}

    @media (prefers-color-scheme: dark) {{
      :root {{
        --bg: #1d1712;
        --bg-strong: #241c16;
        --surface: rgba(39, 31, 24, 0.84);
        --surface-strong: rgba(46, 36, 29, 0.92);
        --ink: #f7efe3;
        --muted: #d1c1ad;
        --line: rgba(255, 241, 224, 0.12);
        --shadow: 0 24px 80px rgba(0, 0, 0, 0.28);
        --hero: linear-gradient(135deg, rgba(182, 129, 66, 0.22), rgba(72, 89, 105, 0.16));
      }}

      body {{
        background:
          radial-gradient(circle at top left, rgba(157, 111, 58, 0.2), transparent 28%),
          radial-gradient(circle at top right, rgba(84, 104, 128, 0.16), transparent 24%),
          linear-gradient(180deg, #18130f 0%, #231b15 100%);
      }}

      .card-shell,
      .panel,
      .meta-card,
      .year-header,
      .search-box,
      .footnote,
      .empty-state {{
        border-color: rgba(255, 241, 224, 0.08);
      }}

      .year-pill,
      .legend-chip,
      .links-block a {{
        background: rgba(45, 36, 28, 0.9);
        border-color: rgba(255, 241, 224, 0.08);
      }}

      .date-tag {{
        background: rgba(255, 241, 224, 0.08);
      }}

      .search-box input {{
        color: var(--ink);
      }}
    }}
  </style>
</head>
<body>
  <main class="page">
    <section class="hero">
      <div class="hero-grid">
        <div class="hero-copy">
          <p class="eyebrow">Editorial Timeline</p>
          <h1>EigenPhi 里程碑时间线</h1>
          <p>
            把 Excel 里的阶段事件压成一条可读的叙事曲线：从 2021 年成立与融资，
            到 2022-2024 年产品与研究报告密集发布，再到 2025-2026 年市场扩张与研究延续。
          </p>
        </div>
        <div class="hero-meta">
          <div class="meta-card">
            <strong>{len(events)}</strong>
            <span>有效里程碑事件</span>
          </div>
          <div class="meta-card">
            <strong>{first_event['year']} → {last_event['year']}</strong>
            <span>时间跨度 {last_event['year'] - first_event['year'] + 1} 年</span>
          </div>
        </div>
      </div>
    </section>

    <section class="info-grid">
      <article class="panel">
        <p class="eyebrow">Snapshot</p>
        <h2>总体概览</h2>
        <div class="stat-row">
          <div class="stat-card">
            <strong>{year_counts.most_common(1)[0][0]}</strong>
            <span>事件最密集年份，共 {year_counts.most_common(1)[0][1]} 条</span>
          </div>
          <div class="stat-card">
            <strong>{category_counts.most_common(1)[0][0]}</strong>
            <span>占比最高类别，共 {category_counts.most_common(1)[0][1]} 条</span>
          </div>
          <div class="stat-card">
            <strong>{sum(1 for item in events if item['links'])}</strong>
            <span>带来源链接的事件</span>
          </div>
        </div>
        <div class="footnote">
          数据源：<code>{html.escape(str(source_path))}</code><br>
          生成时间：<code>{generated_at}</code><br>
          说明：时间字段按中文日期字符串归一化排序；形如 <code>2022年 6-7月</code> 的项按区间起始月排序并保留范围标记。
        </div>
      </article>

      <aside class="panel">
        <p class="eyebrow">Navigate</p>
        <h2>年份与类别</h2>
        <div class="year-nav">
          {year_nav(year_counts)}
        </div>
        <div style="height: 16px;"></div>
        <div class="legend" id="categoryFilters">
          {category_badges(category_counts)}
        </div>
      </aside>
    </section>

    <section class="panel" style="margin-top: 20px;">
      <p class="eyebrow">Explorer</p>
      <h2>筛选与搜索</h2>
      <div class="controls">
        <label class="search-box" aria-label="搜索事件">
          <span>Search</span>
          <input id="searchInput" type="search" placeholder="按事件、战略意义或关键词过滤">
        </label>
        <button class="legend-chip filter-chip is-active" type="button" id="resetFilters"
          style="--chip-accent: #51606b; --chip-soft: #e4eaee;">
          <span class="legend-dot"></span>
          <span>重置筛选</span>
        </button>
      </div>
    </section>

    <section class="timeline-wrap" id="timelineRoot">
      {event_cards(events)}
    </section>

    <div class="empty-state" id="emptyState">
      当前筛选条件下没有匹配事件，重置类别或清空搜索词后会恢复显示。
    </div>
  </main>

  <script id="timeline-data" type="application/json">{stats_json}</script>
  <script>
    const chips = [...document.querySelectorAll('#categoryFilters .filter-chip')];
    const cards = [...document.querySelectorAll('.timeline-card')];
    const searchInput = document.getElementById('searchInput');
    const resetButton = document.getElementById('resetFilters');
    const emptyState = document.getElementById('emptyState');
    const yearBlocks = [...document.querySelectorAll('.year-block')];
    const activeCategories = new Set(chips.map((chip) => chip.dataset.filter));

    function syncChipState() {{
      chips.forEach((chip) => {{
        const active = activeCategories.has(chip.dataset.filter);
        chip.classList.toggle('is-active', active);
        chip.classList.toggle('is-muted', !active);
      }});
    }}

    function updateTimeline() {{
      const query = searchInput.value.trim().toLowerCase();
      let visibleCount = 0;

      cards.forEach((card) => {{
        const matchesCategory = activeCategories.has(card.dataset.category);
        const matchesQuery = !query || card.dataset.search.includes(query);
        const show = matchesCategory && matchesQuery;
        card.hidden = !show;
        if (show) visibleCount += 1;
      }});

      yearBlocks.forEach((block) => {{
        const anyVisible = [...block.querySelectorAll('.timeline-card')].some((card) => !card.hidden);
        block.hidden = !anyVisible;
      }});

      emptyState.classList.toggle('is-visible', visibleCount === 0);
    }}

    chips.forEach((chip) => {{
      chip.addEventListener('click', () => {{
        const key = chip.dataset.filter;
        if (activeCategories.has(key) && activeCategories.size > 1) {{
          activeCategories.delete(key);
        }} else {{
          activeCategories.add(key);
        }}
        if (activeCategories.size === 0) {{
          chips.forEach((entry) => activeCategories.add(entry.dataset.filter));
        }}
        syncChipState();
        updateTimeline();
      }});
    }});

    resetButton.addEventListener('click', () => {{
      activeCategories.clear();
      chips.forEach((chip) => activeCategories.add(chip.dataset.filter));
      searchInput.value = '';
      syncChipState();
      updateTimeline();
    }});

    searchInput.addEventListener('input', updateTimeline);
    syncChipState();
    updateTimeline();
  </script>
</body>
</html>
"""


def main() -> None:
    parser = argparse.ArgumentParser(description="Render a visual HTML timeline from the EigenPhi milestones workbook.")
    parser.add_argument("input", nargs="?", default="/Users/kezheng/Downloads/EigenPhi Milestones.xlsx")
    parser.add_argument("output", nargs="?", default="tasks/eigenphi-milestones-timeline.html")
    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    events = load_events(input_path)
    output_path.write_text(render_html(events, input_path), encoding="utf-8")
    print(output_path)


if __name__ == "__main__":
    main()
